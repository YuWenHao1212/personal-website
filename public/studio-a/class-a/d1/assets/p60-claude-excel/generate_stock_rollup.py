"""
Generate F 組 B3 climax hands-on 資料包: 教育事業部 週進銷存（W27）
For STUDIO A 階段 3 F 組 D1 B3「用它產出」demo + 學員 hands-on

輸出「多來源、格式不一」的資料包（不是一個乾淨 xlsx）—— 對準 53 痛點：
  「要併好幾個來源：EPB 進出貨、品項主檔、標案、季度達成 —— 各是一個檔、格式不一」

四個來源、兩種格式（各有各的 owner，資訊分散）:
  1. EPB_進出貨明細_2026-W27.csv   — CSV · EPB 系統匯出 · 逐筆 ~200+ 筆（含退貨/作廢/調撥雜訊要 filter）
  2. 品項主檔_成本庫存.xlsx        — XLSX · 採購/財會維護 · 10 品項（售價/成本/庫存/近4週均）
  3. 標案進度_2026Q2.xlsx          — XLSX · 業務維護 · 12 案
  4. 季度達成_產品線.csv           — CSV · 業績系統匯出 · 6 產品線（季目標 vs 累計達成率）
  → 打包成 STUDIO_A_W27_資料包.zip（學員丟進 vault Inbox，Claude 讀 4 檔 2 格式、彙整）

為什麼這樣設計:
- 逐筆明細 ~200+ 筆 + 4 來源 2 格式 → show 得出 Claude「讀多來源、彙總、找洞察」的能力
  （10 行一個 xlsx 人眼就看完、沒說服力）。
- 明細正常出貨 sum 回品項主檔埋的總量 → 彙總後 WOI＋毛利 aha 不變。

埋 aha（彙總後，deterministic）:
- iPad A/B = 出貨營收榜首，但 WOI < 2 週（快斷貨、擋校A交機）＋ 毛利 8~10%（最薄）
- 週邊配件 ＋ MDM 授權 = 毛利率 38~62%、WOI 健康 → 真正利潤引擎
- 充電車 E = 積壓品 WOI ~17 週

脫敏：學校「校 A~」、品項通稱、原廠「原廠 X/Y/Z」、業務英文名。
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import os
import csv
import zipfile
import random

rng = random.Random(42)  # reproducible
HERE = os.path.dirname(os.path.abspath(__file__))

SCHOOLS = ["校 A", "校 C", "校 E", "校 H", "校 K", "校 P", "校 W", "校 T"]
STORES = ["門市甲", "門市乙", "門市丙"]
VENDORS = ["原廠 X", "原廠 Y", "原廠 Z"]
WEEK_START = date(2026, 6, 29)   # W27 週一
EDU_CODES = ["EDU-%03d" % i for i in range(1, 13)]

# 品項主檔 — 埋 aha 的相對輪廓（本週出貨/進貨 = 要 explode 成明細的總量）
# 品項, 類目, 教育售價, 毛利率, 期末庫存, 近4週均出貨, 本週進貨, 本週出貨
MASTER = [
    ("iPad 教育版 A",    "平板載具",  11900, 0.08,  210, 150,  60, 168),
    ("iPad 教育版 B",    "平板載具",  15900, 0.10,  150,  88,  40,  96),
    ("MacBook 教育版 C", "筆電",       36900, 0.06,   44,  20,  10,  22),
    ("觸控筆 H",         "週邊配件",   2790, 0.38, 1300, 220, 300, 240),
    ("平板保護殼 D",     "週邊配件",    990, 0.48, 1700, 300, 350, 310),
    ("充電車 E",         "週邊配件",  28900, 0.22,  120,   7,   0,   6),
    ("AP 無線基地台 F",  "網路裝置",   4590, 0.25,  240,  55,  60,  58),
    ("MDM 授權 G",       "MDM 授權",    690, 0.62, None, None, None, 520),
    ("螢幕 I",           "週邊配件",   5990, 0.20,  260,  40,  20,  34),
    ("轉接器 J",         "週邊配件",    890, 0.45,  900, 175, 200, 180),
]

DASH = "—"
_serial = [0]


def bill_no(prefix):
    _serial[0] += 1
    return "%s%06d" % (prefix, _serial[0])


def wk_day():
    return WEEK_START + timedelta(days=rng.randint(0, 5))


def split_qty(total, lo, hi):
    parts, left = [], total
    while left > 0:
        b = min(rng.randint(lo, hi), left)
        parts.append(b)
        left -= b
    return parts


def explode_out(name, cat, price, total):
    rows = []
    if not total:
        return rows
    chosen = rng.sample(SCHOOLS, rng.randint(5, 8))
    w = [rng.random() + 0.25 for _ in chosen]
    ws = sum(w)
    alloc = [max(1, round(total * x / ws)) for x in w]
    alloc[alloc.index(max(alloc))] += total - sum(alloc)
    if price <= 1500:      blo, bhi = 15, 40
    elif price <= 6000:    blo, bhi = 6, 16
    elif price <= 15000:   blo, bhi = 3, 8
    else:                  blo, bhi = 1, 3
    for school, a in zip(chosen, alloc):
        if a <= 0:
            continue
        for q in split_qty(a, blo, bhi):
            rows.append([wk_day().isoformat(), bill_no("SO"), "出貨", school, "學校",
                         name, cat, q, price, q * price, rng.choice(EDU_CODES), "正常"])
    return rows


def explode_in(name, cat, cost, total):
    rows = []
    if not total:
        return rows
    for q in split_qty(total, max(1, total // 3), max(2, total // 2)):
        rows.append([wk_day().isoformat(), bill_no("PO"), "進貨", rng.choice(VENDORS), "原廠",
                     name, cat, q, cost, q * cost, DASH, "正常"])
    return rows


# ---- xlsx styling ----
HFONT = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
HFILL = PatternFill(start_color="1D1D1F", end_color="1D1D1F", fill_type="solid")
BFONT = Font(name="微軟正黑體", size=10)
THIN = Border(*[Side(style="thin", color="D0D0D0")] * 4)


def save_xlsx(path, title, header, data, money_cols=(), pct_cols=(), date_cols=()):
    wb = Workbook(); ws = wb.active; ws.title = title
    ws.append(header)
    for row in data:
        ws.append(row)
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HFONT; cell.fill = HFILL
        cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = THIN
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c); cell.font = BFONT; cell.border = THIN
        for c in money_cols: ws.cell(row=r, column=c).number_format = '"NT$"#,##0'
        for c in pct_cols:   ws.cell(row=r, column=c).number_format = '0%'
        for c in date_cols:  ws.cell(row=r, column=c).number_format = "YYYY-MM-DD"
    for c in range(1, ws.max_column + 1):
        L = get_column_letter(c)
        m = max((sum(2 if ord(ch) > 127 else 1 for ch in str(ws.cell(row=r, column=c).value))
                 for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=c).value is not None), default=8)
        ws.column_dimensions[L].width = min(max(m + 2, 10), 30)
    wb.save(path)


def save_csv(path, header, data):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(data)


def main():
    files = []

    # ===== 1. EPB 進出貨明細（CSV，逐筆 ~200+）=====
    detail = []
    net_out = {}
    for (name, cat, price, margin, stock, avg4, buy, sell) in MASTER:
        cost = round(price * (1 - margin))
        o = explode_out(name, cat, price, sell)
        detail += o
        net_out[name] = sum(r[7] for r in o)
        detail += explode_in(name, cat, cost, buy)
    for _ in range(18):  # 調撥雜訊
        name, cat, price, *_ = rng.choice(MASTER)
        detail.append([wk_day().isoformat(), bill_no("TR"), "調撥", rng.choice(STORES), "門市",
                       name, cat, rng.randint(2, 20), DASH, DASH, DASH, "調撥中"])
    for _ in range(14):  # 退貨/作廢雜訊（要 filter）
        name, cat, price, *_ = rng.choice(MASTER)
        q = rng.randint(1, 6)
        typ, st = rng.choice([("退貨", "退貨"), ("出貨", "作廢")])
        detail.append([wk_day().isoformat(), bill_no("RMA" if typ == "退貨" else "SO"), typ,
                       rng.choice(SCHOOLS), "學校", name, cat, q, price, q * price,
                       rng.choice(EDU_CODES), st])
    detail.sort(key=lambda r: (r[0], r[1]))
    p1 = os.path.join(HERE, "EPB_進出貨明細_2026-W27.csv")
    save_csv(p1, ["日期", "單號", "類型", "對象", "對象類型", "品項", "類目",
                  "數量", "單價", "金額", "教育專案", "狀態"], detail)
    files.append(p1)

    # ===== 2. 品項主檔（XLSX，採購/財會）=====
    master_rows = []
    for (name, cat, price, margin, stock, avg4, buy, sell) in MASTER:
        cost = round(price * (1 - margin))
        master_rows.append([name, cat, price, cost,
                            stock if stock is not None else DASH,
                            avg4 if avg4 is not None else DASH])
    p2 = os.path.join(HERE, "品項主檔_成本庫存.xlsx")
    save_xlsx(p2, "品項主檔", ["品項", "類目", "教育售價", "進貨成本", "期末庫存", "近4週均出貨"],
              master_rows, money_cols=(3, 4))
    files.append(p2)

    # ===== 3. 標案進度（XLSX，業務）=====
    REPS = ["Edick", "Rabio", "Jim", "Asa", "Gariano", "Eddy", "KC", "Wesley", "Steve", "Max", "Edick", "Jim"]
    BIDS = [
        ("全校 iPad 教學載具", "校 A", 1680000, "交機中", date(2026, 7, 10), 0.85),
        ("教師 MacBook 汰換", "校 C", 890000, "評選", date(2026, 7, 18), 0.60),
        ("電腦教室更新", "校 E", 2400000, "報價", date(2026, 7, 24), 0.35),
        ("行動學習裝置擴充", "校 H", 560000, "已決標", date(2026, 6, 20), 1.00),
        ("校務 MDM 平台", "校 K", 1200000, "評選", date(2026, 6, 26), 0.45),
        ("圖書館自助借還", "校 P", 320000, "決標前", date(2026, 7, 15), 0.78),
        ("英語情境教室平板", "校 W", 780000, "初訪", date(2026, 8, 5), 0.20),
        ("特教輔具平板", "校 T", 450000, "已決標", date(2026, 6, 18), 1.00),
        ("智慧教室單槍汰換", "校 C", 640000, "報價", date(2026, 7, 30), 0.30),
        ("行政 iPad 擴充", "校 K", 380000, "決標前", date(2026, 7, 8), 0.82),
        ("圖書館平板借閱", "校 A", 290000, "評選", date(2026, 7, 12), 0.55),
        ("充電車補充採購", "校 H", 210000, "初訪", date(2026, 8, 12), 0.15),
    ]
    bid_rows = [[nm, sch, amt, stg, dd, rate, rep]
                for (nm, sch, amt, stg, dd, rate), rep in zip(BIDS, REPS)]
    p3 = os.path.join(HERE, "標案進度_2026Q2.xlsx")
    save_xlsx(p3, "標案進度", ["標案", "學校", "金額", "階段", "預計決標日", "達成率", "負責人"],
              bid_rows, money_cols=(3,), pct_cols=(6,), date_cols=(5,))
    files.append(p3)

    # ===== 4. 季度達成 各產品線（CSV，業績系統匯出）=====
    REPORT = [("平板載具", "台", 12000, 9600), ("筆電", "台", 1500, 1020),
              ("週邊配件", "件", 26000, 23920), ("網路裝置", "台", 2400, 1776),
              ("MDM 授權", "套", 8000, 6880), ("維修服務", "件", 900, 612)]
    rep_rows = [[line, unit, goal, cum, "%.0f%%" % (cum / goal * 100)] for (line, unit, goal, cum) in REPORT]
    p4 = os.path.join(HERE, "季度達成_產品線.csv")
    save_csv(p4, ["產品線", "單位", "本季目標", "本季累計", "達成率"], rep_rows)
    files.append(p4)

    # ===== 打包 zip =====
    zip_path = os.path.join(HERE, "STUDIO_A_W27_資料包.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for fp in files:
            z.write(fp, arcname=os.path.basename(fp))

    # ===== 健檢 key =====
    n_out = sum(1 for r in detail if r[2] == "出貨" and r[11] == "正常")
    n_noise = sum(1 for r in detail if r[11] in ("退貨", "作廢", "調撥中"))
    print("✓ 資料包（4 來源 · 2 格式）→", zip_path)
    for fp in files:
        print("   -", os.path.basename(fp))
    print("\n【W27 健檢 key】")
    print("  EPB 明細 %d 筆（正常出貨 %d · 退貨/作廢/調撥雜訊 %d 要 filter）· 主檔 10 · 標案 12 · 季度達成 6"
          % (len(detail), n_out, n_noise))

    cat_rev, cat_prof = {}, {}
    tot_rev = tot_prof = 0
    per = {}
    m_by = {m[0]: m for m in MASTER}
    for name, out in net_out.items():
        _, cat, price, margin, stock, avg4, buy, sell = m_by[name]
        rev = out * price; prof = int(round(rev * margin))
        woi = round(stock / avg4, 1) if stock and avg4 else None
        per[name] = (cat, out, rev, margin, prof, woi)
        tot_rev += rev; tot_prof += prof
        cat_rev[cat] = cat_rev.get(cat, 0) + rev
        cat_prof[cat] = cat_prof.get(cat, 0) + prof
    print("  本週出貨營收 NT$%s · 毛利 NT$%s（綜合 %.1f%%）" % (f"{tot_rev:,}", f"{tot_prof:,}", tot_prof / tot_rev * 100))
    rr = sorted(cat_rev, key=cat_rev.get, reverse=True)
    pr = sorted(cat_prof, key=cat_prof.get, reverse=True)
    print("  營收榜:", " > ".join("%s %.0f%%" % (c, cat_rev[c] / tot_rev * 100) for c in rr))
    print("  毛利榜:", " > ".join("%s %.0f%%" % (c, cat_prof[c] / tot_prof * 100) for c in pr))
    print("  WOI:")
    for name, (cat, out, rev, margin, prof, woi) in sorted(per.items(), key=lambda x: (x[1][5] is None, x[1][5])):
        if woi is None:
            print("    %-14s WOI —（授權無庫存）· 出貨 %d" % (name, out)); continue
        flag = "🔴 見底" if woi < 2 else ("🟠 積壓" if woi > 15 else "🟢")
        print("    %-14s WOI %4.1f · 出貨 %d · 毛利 %.0f%%  %s" % (name, woi, out, margin * 100, flag))
    print("  ＊淨出貨 sum 回主檔埋的總量 → 彙總後 aha 不變")


if __name__ == "__main__":
    main()
