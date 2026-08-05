"""
Generate F 組 Track 1 fake Excel: 校園客戶標案彙整表
For STUDIO A 階段 3 F 組 D1 hands-on (跑場景 · Track 1)

Output: ../../../../2026-06_校園客戶標案彙整表.xlsx (class-f 根目錄)
- Sheet 1: 客戶主檔    (客戶編號/學校代號/類型/區域/負責業務/合作狀態/最後拜訪日)
- Sheet 2: 標案進度    (標案編號/客戶編號/標案名稱/金額/階段/預計決標日/狀態)
- Sheet 3: 拜訪記錄    (客戶編號/拜訪日/業務/摘要/下一步)

設計理由（對準 F 組痛點「多個數據表格要交叉確認、資訊分散難整理」）:
- 三張表要用「客戶編號」join 才看得出全貌 → break 學員的 VLOOKUP、show Claude 的 join 韌性
- 「本週待跟進客戶」= 標案在關鍵階段(評選/決標) 但最後拜訪 > 14 天沒接觸 → 要跨「標案」+「拜訪」兩表算
- 「標案進度異常」= 預計決標日已過還沒決標 / 金額異常 → 要跨「標案」自檢 + 抓髒資料
- 種 5 筆髒資料 break 直接 join:
    G1. 孤兒標案：客戶編號 C099 在客戶主檔查無 (join 漏)
    G2. 金額存成文字「1,200,000」(含逗號、不能加總)
    G3. 重複客戶：同一所學校兩個編號 (C005 / C025)、不同業務登記 → 跨表才發現
    G4. 預計決標日 typo 2027 年 (門市/業務 key 錯、造成「未來很遠」假象)
    G5. 拜訪記錄客戶編號全形「Ｃ003」對不上主檔「C003」(全形/半形 join 破)

基準日（「本週」）= 2026-06-26（週五）。
脫敏：學校用「校 A ~ 校 Z」代號、業務用英文名、不用真實校名 / 客戶 / 員工資料。
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import os
import random

random.seed(42)  # reproducible

BASE_DAY = date(2026, 6, 26)   # 「本週」基準（週五）
STALE_DAYS = 14                # 超過 14 天沒拜訪 = 待跟進門檻

REPS = ["Edick", "Rabio", "Jim", "Javon", "Asa", "Gariano", "Eddy", "KC", "Wesley", "Steve", "Max"]
REGIONS = ["北一區", "北二區", "桃竹區", "中區", "南區"]
SCHOOL_TYPES = ["國立大學", "私立大學", "高中職", "國中小", "私立完全中學"]
STAGES = ["初訪", "報價", "評選", "決標前", "已決標", "交機中"]
KEY_STAGES = {"評選", "決標前"}          # 關鍵階段 → 久未接觸就要跟進
DONE_STATES = {"已決標", "交機完成"}     # 這些狀態就算過了決標日也正常

# ---- 30 客戶主檔（校 A ~ 校 Z + 私校）----
SCHOOL_CODES = [
    "校 A（市立高中）", "校 B（國立科大）", "校 C（私立大學）", "校 D（市立國中）",
    "校 E（國立大學）", "校 F（私立高中）", "校 G（縣立國小）", "校 H（國立高中）",
    "校 I（私立完中）", "校 J（市立高職）", "校 K（國立大學）", "校 L（私立科大）",
    "校 M（市立國小）", "校 N（國立高中）", "校 O（私立大學）", "校 P（縣立國中）",
    "校 Q（國立科大）", "校 R（市立高中）", "校 S（私立完中）", "校 T（國立大學）",
    "校 U（市立國中）", "校 V（私立高職）", "校 W（國立高中）", "校 X（縣立國小）",
    "校 Y（私立大學）", "校 Z（國立科大）", "校 AA（市立高中）", "校 AB（私立國中）",
    "校 AC（國立大學）", "校 AD（市立高職）",
]


# ==== 髒資料 / 答案 預埋（確定性）====
# 待跟進客戶：關鍵階段 + 最後拜訪超過 14 天。以「客戶編號 idx(0-based)」指定。
STALE_FOLLOWUP_IDX = [2, 7, 10, 15, 22]   # C003 C008 C011 C016 C023
# 進度異常：預計決標日已過、狀態未決標。以「標案 index」指定（見下 BID 生成）。
# 金額文字髒資料客戶：
AMOUNT_TEXT_BID = 6           # 第 7 筆標案金額存成文字「1,200,000」(G2)
ORPHAN_BID_IDX = 8            # 第 9 筆標案設孤兒（避開 stale-followup 0-4）
ORPHAN_BID_CUST = "C099"      # G1 孤兒標案客戶編號
DUP_SCHOOL_PRIMARY = 4        # C005（校 E 國立大學）
DUP_SCHOOL_SECOND = 24        # C025 = 校 E 重複登記（G3）
DATE_TYPO_BID = 11            # 第 12 筆標案 預計決標日 typo 2027（G4）
FULLWIDTH_VISIT_IDX = 2       # 拜訪記錄用全形「Ｃ003」對不上（G5）

# Styling
HEADER_FONT = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="3A281E", end_color="3A281E", fill_type="solid")
BODY_FONT = Font(name="微軟正黑體", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER


def style_body(ws):
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=row, column=col)
            c.font = BODY_FONT
            c.border = THIN_BORDER


def autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                length = sum(2 if ord(ch) > 127 else 1 for ch in str(v))
                max_len = max(max_len, length)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 34)


def cid(idx):
    return "C%03d" % (idx + 1)


def main():
    wb = Workbook()

    # ===== Sheet 1: 客戶主檔 =====
    ws1 = wb.active
    ws1.title = "客戶主檔"
    ws1.append(["客戶編號", "學校代號", "類型", "區域", "負責業務", "合作狀態", "最後拜訪日"])
    customers = []
    for idx, school in enumerate(SCHOOL_CODES):
        stype = SCHOOL_TYPES[idx % len(SCHOOL_TYPES)]
        region = REGIONS[idx % len(REGIONS)]
        rep = REPS[idx % len(REPS)]
        status = random.choice(["洽談中", "合作中", "合作中", "已結案"])
        # 最後拜訪日：待跟進 idx 塞遠日、其餘塞近日
        if idx in STALE_FOLLOWUP_IDX:
            last_visit = BASE_DAY - timedelta(days=random.randint(18, 40))
        else:
            last_visit = BASE_DAY - timedelta(days=random.randint(1, 12))
        # G3 重複客戶：C025 = 校 E 重複（覆蓋學校代號 + 換業務）
        if idx == DUP_SCHOOL_SECOND:
            school = SCHOOL_CODES[DUP_SCHOOL_PRIMARY]  # 同 校 E
            rep = "Wesley"  # 不同業務登記
            status = "洽談中"
        row = [cid(idx), school, stype, region, rep, status, last_visit]
        customers.append(row)
        ws1.append(row)
    style_header(ws1, 7)
    style_body(ws1)
    autosize_columns(ws1)
    for r in range(2, ws1.max_row + 1):
        ws1.cell(row=r, column=7).number_format = "YYYY-MM-DD"

    # ===== Sheet 2: 標案進度 =====
    ws2 = wb.create_sheet("標案進度")
    ws2.append(["標案編號", "客戶編號", "標案名稱", "金額", "階段", "預計決標日", "狀態"])
    BID_NAMES = [
        "全校 iPad 教學載具採購", "教師 MacBook 汰換案", "電腦教室更新標案",
        "行動學習裝置擴充", "校務 MDM 管理平台", "圖書館自助借還裝置",
        "英語情境教室平板案", "體育館數位看板", "宿舍網路裝置更新",
        "特教輔具平板採購", "行政 iPad 派發案", "校園導覽 App 載具",
    ]
    bids = []
    planted_abnormal = []   # 進度異常答案（標案編號）
    for i in range(40):
        # 客戶編號：待跟進客戶(0-4)強制關鍵階段標案；第 9 筆設孤兒；其餘隨機
        if i < len(STALE_FOLLOWUP_IDX):
            cust = cid(STALE_FOLLOWUP_IDX[i])
            stage = random.choice(list(KEY_STAGES))
        elif i == ORPHAN_BID_IDX:
            cust = ORPHAN_BID_CUST                      # G1 孤兒（客戶主檔查無）
            stage = random.choice(STAGES)
        else:
            cust = cid(random.randint(0, len(SCHOOL_CODES) - 1))
            stage = random.choice(STAGES)
        name = BID_NAMES[i % len(BID_NAMES)]
        amount = random.choice([180000, 320000, 560000, 890000, 1200000, 1680000, 2400000])
        # 預計決標日
        exp = BASE_DAY + timedelta(days=random.randint(-30, 45))
        status = "進行中"
        if stage in DONE_STATES or stage == "交機中":
            status = "已決標"
        # G4 日期 typo 2027
        if i == DATE_TYPO_BID:
            exp = date(2027, exp.month, exp.day)
        # 進度異常預埋：預計決標日已過 + 未決標
        if 20 <= i <= 23:   # 塞 4 筆明確異常
            exp = BASE_DAY - timedelta(days=random.randint(5, 20))
            stage = "評選"
            status = "進行中"
            planted_abnormal.append("F%03d" % (i + 1))
        bid_no = "F%03d" % (i + 1)
        # G2 金額文字
        amount_cell = "1,200,000" if i == AMOUNT_TEXT_BID else amount
        row = [bid_no, cust, name, amount_cell, stage, exp, status]
        bids.append(row)
        ws2.append(row)
    style_header(ws2, 7)
    style_body(ws2)
    autosize_columns(ws2)
    for r in range(2, ws2.max_row + 1):
        v = ws2.cell(row=r, column=4).value
        if isinstance(v, (int, float)):
            ws2.cell(row=r, column=4).number_format = '"NT$"#,##0'
        ws2.cell(row=r, column=6).number_format = "YYYY-MM-DD"

    # ===== Sheet 3: 拜訪記錄 =====
    ws3 = wb.create_sheet("拜訪記錄")
    ws3.append(["客戶編號", "拜訪日", "業務", "摘要", "下一步"])
    SUMMARIES = [
        "初步需求訪談、確認採購時程", "報價說明、對方要內部簽核", "demo 教學載具、老師反應佳",
        "評選簡報、競爭對手也進場", "確認規格與保固條款", "跟進決標時程、對方預算待核",
        "交機安裝協調", "售後教育訓練排期", "追加裝置需求討論",
    ]
    NEXT = ["下週回報價", "等對方簽核", "約 demo", "補送規格書", "跟進決標", "排交機", "約教育訓練"]
    for i in range(52):
        # 大多對得上客戶主檔；待跟進客戶的拜訪日刻意設遠（呼應主檔）
        c_idx = random.randint(0, len(SCHOOL_CODES) - 1)
        cust = cid(c_idx)
        visit = BASE_DAY - timedelta(days=random.randint(1, 45))
        # G5 全形客戶編號對不上（第 3 筆）
        if i == 3:
            cust = "Ｃ003"
            visit = BASE_DAY - timedelta(days=6)
        rep = REPS[c_idx % len(REPS)]
        row = [cust, visit, rep, random.choice(SUMMARIES), random.choice(NEXT)]
        ws3.append(row)
    style_header(ws3, 5)
    style_body(ws3)
    autosize_columns(ws3)
    for r in range(2, ws3.max_row + 1):
        ws3.cell(row=r, column=2).number_format = "YYYY-MM-DD"

    # Output → class-f 根目錄（4 層上）
    root = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "..", ".."))
    output_path = os.path.join(root, "2026-06_校園客戶標案彙整表.xlsx")
    wb.save(output_path)
    print("✓ Generated:", output_path)

    # ===== 答案 key =====
    print("\n【基準日】%s（週五）· 待跟進門檻 = 最後拜訪 > %d 天" % (BASE_DAY, STALE_DAYS))
    print("\n[A] 本週待跟進客戶（關鍵階段 + 久未接觸）應標 %d 筆：" % len(STALE_FOLLOWUP_IDX))
    for idx in STALE_FOLLOWUP_IDX:
        c = customers[idx]
        gap = (BASE_DAY - c[6]).days
        print("  %s %s | 業務 %s | 最後拜訪 %s（%d 天前）" % (c[0], c[1], c[4], c[6], gap))
    print("\n[B] 標案進度異常（預計決標日已過、未決標）應標 %d 筆：" % len(planted_abnormal))
    for i, b in enumerate(bids):
        if b[0] in planted_abnormal:
            print("  %s 客戶 %s | %s | 預計決標 %s（已過）| 階段 %s" % (b[0], b[1], b[2], b[5], b[4]))
    print("\n[C] 髒資料預埋（跨表交叉才抓得到）：")
    print("  G1 孤兒標案 : F%03d 客戶編號 %s 在客戶主檔查無" % (ORPHAN_BID_IDX + 1, ORPHAN_BID_CUST))
    print("  G2 金額文字 : F%03d 金額存成文字「1,200,000」(含逗號、不能加總)" % (AMOUNT_TEXT_BID + 1))
    print("  G3 重複客戶 : C005 與 C025 同為「%s」、不同業務登記" % SCHOOL_CODES[DUP_SCHOOL_PRIMARY])
    print("  G4 日期 typo: F%03d 預計決標日 = 2027 年 (應為 2026)" % (DATE_TYPO_BID + 1))
    print("  G5 全形編號 : 拜訪記錄「Ｃ003」全形、對不上主檔「C003」")


if __name__ == "__main__":
    main()
