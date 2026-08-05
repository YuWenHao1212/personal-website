"""
Generate F 組 Track 2 fake Excel: 進出貨原廠報備核對表
For STUDIO A 階段 3 F 組 D1 hands-on (跑場景 · Track 2)

Output: ../../../../2026-06_進出貨原廠報備核對表.xlsx (class-f 根目錄)
- Sheet 1: 進出貨明細   (單號/日期/類型/品項類目/型號/數量/對象/教育專案代號/狀態)
- Sheet 2: 核對規則     (12 條 + 教育專案授權清單)

設計理由（對準 F 組痛點「原廠報備跨表拼、門市/原廠資料有誤要特別撈」）:
- 規則寫進第 2 分頁 → 學員叫 Claude「連核對規則分頁一起讀」即可（mirror B 組 Kidd v2.1）
- 種 13 筆疑慮 / 12 型、分兩軸:
    [A] 報備 / 規則（原廠報備本職）
    [B] 資料有誤（門市/系統 key 錯 → 要特別撈）
- 兩個需要「跨列 / 跨欄」比對的 check（原廠報備數量加總、同單號重複）最容易漏 → 追問教學點

脫敏：學校用代號、型號用通稱、原廠用「原廠 X/Y/Z」、不用真實 SKU / 校名 / 員工。
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import os
import random

random.seed(42)  # reproducible

MONTH = (2026, 6)   # 當月：進出貨日期須落在 2026-06

TYPES = ["進貨", "出貨", "調撥"]
CATEGORIES = ["平板載具", "筆電", "MDM 授權", "週邊配件", "網路裝置"]
MODELS = [
    "iPad 教育版 A", "iPad 教育版 B", "MacBook 教育版 C", "平板保護殼 D",
    "充電車 E", "AP 無線基地台 F", "MDM 授權 G", "觸控筆 H", "螢幕 I", "轉接器 J",
]
VENDORS = ["原廠 X", "原廠 Y", "原廠 Z"]
TARGETS_SCHOOL = ["校 A", "校 C", "校 E", "校 H", "校 K", "校 P", "校 W"]
TARGETS_STORE = ["門市甲", "門市乙", "門市丙"]
EDU_CODES = ["EDU-%03d" % i for i in range(1, 21)]   # 授權清單 EDU-001 ~ EDU-020

# ==== 髒資料 / 疑慮 預埋（確定性，以「明細列 index」= 0-based 指定）====
# [A] 報備 / 規則
Q_QTY_OVER_NO_REPORT = 3      # 出貨 > 50 台但無原廠報備編號（規則 1）
Q_EDU_UNAUTH = 9              # 教育專案代號 EDU-099 不在授權清單（規則 2）
Q_SCHOOL_NO_EDU = 14          # 出貨對象=學校 但教育專案代號留空（規則 7）
Q_REPORT_MISMATCH_A = 20      # 原廠 X 報備數量與出貨加總對不上（規則 3）跨列
Q_REPORT_MISMATCH_B = 21      # （同上、第二列，湊出加總差異）
# [B] 資料有誤
Q_DATE_2027 = 5              # 日期 2027（規則 4）
Q_QTY_NEG = 12              # 數量 -8（規則 5）
Q_QTY_ZERO = 27            # 數量 0（規則 5）
Q_MODEL_TYPO = 18          # 型號全形「ｉPad 教育版 A」對不上型號主檔（規則 6）
Q_DUP_ORDER = 33           # 同單號重複（規則 9）→ 與第 34 列同單號
Q_TARGET_BLANK = 24        # 對象欄留空（規則 8/資料完整）
Q_TRANSFER_NO_SRC = 30     # 調撥未註明來源門市（規則 8）

# Styling
HEADER_FONT = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="3A281E", end_color="3A281E", fill_type="solid")
RULE_FILL = PatternFill(start_color="9F462E", end_color="9F462E", fill_type="solid")
BODY_FONT = Font(name="微軟正黑體", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def style_header(ws, ncols, fill=HEADER_FILL):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.font = HEADER_FONT
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER


def style_body(ws):
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=row, column=col)
            c.font = BODY_FONT
            c.border = THIN_BORDER


def autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                length = sum(2 if ord(ch) > 127 else 1 for ch in str(v))
                max_len = max(max_len, length)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 30)


def order_no(i):
    return "IO2026060%03d" % (i + 1)


def main():
    wb = Workbook()

    # ===== Sheet 1: 進出貨明細 =====
    ws1 = wb.active
    ws1.title = "進出貨明細"
    ws1.append(["單號", "日期", "類型", "品項類目", "型號", "數量", "對象", "教育專案代號", "原廠報備編號", "狀態"])

    rows = []
    issues = []   # (單號, 型別, 說明)

    for i in range(60):
        d = date(MONTH[0], MONTH[1], random.randint(1, 26))
        typ = random.choice(TYPES)
        cat = random.choice(CATEGORIES)
        model = random.choice(MODELS)
        qty = random.randint(3, 40)
        vendor = random.choice(VENDORS)
        # 對象
        if typ == "出貨":
            target = random.choice(TARGETS_SCHOOL)
            edu = random.choice(EDU_CODES)
        elif typ == "進貨":
            target = vendor
            edu = ""
        else:  # 調撥
            target = "%s → %s" % (random.choice(TARGETS_STORE), random.choice(TARGETS_STORE))
            edu = ""
        report_no = ""
        if typ == "出貨" and qty > 50:
            report_no = "RPT-%s-%03d" % (vendor[-1], i)
        status = "已報備" if report_no else "待處理"
        ono = order_no(i)

        # ---- 髒資料 / 疑慮覆蓋 ----
        if i == Q_QTY_OVER_NO_REPORT:      # A: 出貨 55 台無報備編號
            typ, qty, target, edu, report_no = "出貨", 55, random.choice(TARGETS_SCHOOL), random.choice(EDU_CODES), ""
            status = "待處理"
            issues.append((ono, "A·規則1", "出貨 55 台 > 50 台上限，但原廠報備編號留空"))
        elif i == Q_EDU_UNAUTH:            # A: 教育專案代號未授權
            typ, target, edu = "出貨", random.choice(TARGETS_SCHOOL), "EDU-099"
            issues.append((ono, "A·規則2", "教育專案代號 EDU-099 不在授權清單（EDU-001~020）"))
        elif i == Q_SCHOOL_NO_EDU:         # A: 出貨學校但 edu 空
            typ, target, edu = "出貨", random.choice(TARGETS_SCHOOL), ""
            issues.append((ono, "A·規則7", "出貨對象為學校，但教育專案代號留空"))
        elif i == Q_REPORT_MISMATCH_A:     # A: 原廠 X 出貨（跨列加總會對不上）
            typ, cat, model, qty, target, edu, vendor = "出貨", "平板載具", "iPad 教育版 A", 30, "校 A", random.choice(EDU_CODES), "原廠 X"
            report_no = "RPT-X-報備 40 台"
            status = "已報備"
            issues.append((ono, "A·規則3", "原廠 X 報備寫 40 台，但本月原廠 X 出貨加總為 50 台（30+20）→ 少報 10 台"))
        elif i == Q_REPORT_MISMATCH_B:     # A: 原廠 X 第二筆出貨（湊 50 加總）
            typ, cat, model, qty, target, edu, vendor = "出貨", "平板載具", "iPad 教育版 A", 20, "校 C", random.choice(EDU_CODES), "原廠 X"
            report_no = ""
            status = "待處理"
            # 不獨立列 issue、屬 A·規則3 的另一半
        elif i == Q_DATE_2027:             # B: 日期 2027
            d = date(2027, 6, 8)
            issues.append((ono, "B·規則4", "日期 2027-06-08 不在當月（2026-06），疑 key 錯"))
        elif i == Q_QTY_NEG:               # B: 數量負
            qty = -8
            issues.append((ono, "B·規則5", "數量 -8 為負數（應 > 0）"))
        elif i == Q_QTY_ZERO:              # B: 數量 0
            qty = 0
            issues.append((ono, "B·規則5", "數量 0（應 > 0）"))
        elif i == Q_MODEL_TYPO:            # B: 型號全形
            model = "ｉPad 教育版 A"
            issues.append((ono, "B·規則6", "型號「ｉPad 教育版 A」全形 i、對不上型號主檔"))
        elif i == Q_TARGET_BLANK:          # B: 對象空
            target = ""
            issues.append((ono, "B·規則8", "對象欄留空（資料不完整）"))
        elif i == Q_TRANSFER_NO_SRC:       # B: 調撥沒來源
            typ, target, edu = "調撥", "門市乙", ""   # 缺「→」來源
            issues.append((ono, "B·規則8", "調撥未註明來源門市（只有目的、缺來源）"))

        row = [ono, d, typ, cat, model, qty, target, edu, report_no, status]
        rows.append(row)
        ws1.append(row)

    # G/Q dup order：第 34 列複製第 33 列單號
    dup_src = rows[Q_DUP_ORDER]
    dup_row = [dup_src[0]] + [
        date(MONTH[0], MONTH[1], 20), "出貨", "筆電", "MacBook 教育版 C", 12,
        random.choice(TARGETS_SCHOOL), random.choice(EDU_CODES), "", "待處理",
    ]
    ws1.append(dup_row)
    issues.append((dup_src[0], "B·規則9", "同一單號 %s 出現兩列（重複）" % dup_src[0]))

    style_header(ws1, 10)
    style_body(ws1)
    autosize_columns(ws1)
    for r in range(2, ws1.max_row + 1):
        ws1.cell(row=r, column=2).number_format = "YYYY-MM-DD"

    # ===== Sheet 2: 核對規則 =====
    ws2 = wb.create_sheet("核對規則")
    ws2.append(["編號", "規則", "說明"])
    RULES = [
        ("規則 1", "出貨單筆數量 > 50 台，須附原廠報備編號", "> 50 台屬大量出貨、原廠要求事前報備；報備編號欄不可空"),
        ("規則 2", "教育專案代號須在授權清單", "見下方授權清單 EDU-001 ~ EDU-020；不在清單 = 未授權專案"),
        ("規則 3", "同一原廠當月出貨加總 = 該原廠報備數量", "跨列加總同原廠出貨、對照報備編號上的數字；對不上 = 漏報/多報"),
        ("規則 4", "進出貨日期須落在當月（2026-06）", "不在當月 = 系統/門市 key 錯"),
        ("規則 5", "數量須為正整數（> 0）", "負數 / 0 = 資料有誤"),
        ("規則 6", "型號須為半形、對得上型號主檔", "全形字 / typo 會讓型號比對失敗"),
        ("規則 7", "出貨對象為學校時，須有教育專案代號", "校園出貨走教育專案、代號不可空"),
        ("規則 8", "調撥須註明來源與目的門市；對象欄不可空", "格式「來源 → 目的」；缺來源或整欄空 = 不完整"),
        ("規則 9", "同一單號不可重複", "重複單號會造成重複出貨 / 報備"),
        ("規則 10", "MDM 授權類不需原廠報備", "授權為數位交付、免報備（避免誤判規則 1）"),
        ("規則 11", "進貨對象須為原廠", "進貨來源記原廠 X/Y/Z"),
        ("規則 12", "已報備狀態須有原廠報備編號", "狀態=已報備 但編號空 = 矛盾"),
    ]
    for r in RULES:
        ws2.append(list(r))
    ws2.append(["", "", ""])
    ws2.append(["授權清單", "教育專案代號", "、".join(EDU_CODES)])
    style_header(ws2, 3, fill=RULE_FILL)
    style_body(ws2)
    autosize_columns(ws2)
    ws2.column_dimensions["C"].width = 60

    # Output → class-f 根目錄（4 層上）
    root = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "..", ".."))
    output_path = os.path.join(root, "2026-06_進出貨原廠報備核對表.xlsx")
    wb.save(output_path)
    print("✓ Generated:", output_path)

    # ===== 答案 key =====
    print("\n【當月】2026-06 · 規則在第 2 分頁（12 條 + 授權清單）")
    print("\n應挑出疑慮 %d 筆：" % len(issues))
    for ono, typ, desc in issues:
        print("  %s [%s] %s" % (ono, typ, desc))
    print("\n※ 最容易漏（需跨列/跨欄）：")
    print("  · 規則3 原廠 X 報備 40 台 vs 出貨加總 50 台（跨兩列 F 出貨相加）")
    print("  · 規則9 同單號重複（跨列比對單號）")


if __name__ == "__main__":
    main()
